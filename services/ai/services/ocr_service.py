"""
Stage 1 OCR service — uses gemma4:31b-cloud (Ollama vision) to scan an
uploaded OPPM form image/PDF and extract its text + field structure.

XLSX path (Google Sheet):
  openpyxl renders the sheet to a PNG image → same gemma4 vision pipeline.

The model reads the image pixels and returns:
  - raw_text  : verbatim text found in the form
  - fields    : dict mapping OPPM field labels → detected values
  - form_rules: structured list describing each field and its location/pattern
"""

import base64
import io
import json
import logging
import re
from typing import Optional

import httpx

from config import get_settings
from infrastructure.llm.base import ProviderUnavailableError
from schemas.ocr_fill import OcrResult

logger = logging.getLogger(__name__)

# ── Model assignment ──────────────────────────────────────────────────────────
# Stage 1 is always gemma4:31b-cloud via the configured Ollama endpoint.
OCR_MODEL_ID = "gemma4:31b-cloud"

# MIME types treated as spreadsheet (skip binary OCR, render via openpyxl)
XLSX_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",  # some backends return this for xlsx
}
XLSX_MAGIC = b"PK\x03\x04"  # zip/xlsx magic bytes

_OCR_PROMPT = """You are an expert OCR and form-analysis engine.

Examine this OPPM (One Page Project Manager) form image carefully.

Your job:
1. Extract ALL visible text, exactly as printed.
2. Identify each form field and its current value (may be empty).
3. Describe the structural rules of this form (header fields, task rows, owner columns, timeline columns, etc.)

Return ONLY valid JSON with this exact structure:
{
  "raw_text": "<all visible text verbatim, newline-separated>",
  "fields": {
    "<field_label>": "<field_value or empty string if blank>"
  },
  "form_rules": {
    "header_fields": ["list of header field names found"],
    "task_row_pattern": "<description of how task rows are structured>",
    "owner_columns": ["list of owner column headers found"],
    "timeline_columns": ["list of week/period column headers found"],
    "notes": "<any other structural observations>"
  }
}

Rules:
- Do NOT add commentary outside the JSON.
- If a field has no visible value, return an empty string "".
- Preserve exact text casing."""


# ── XLSX → PNG renderer ───────────────────────────────────────────────────────

_OPPM_LABEL_MAP = {
    "project name": "project_name",
    "project title": "project_name",
    "project leader": "project_leader",
    "leader": "project_leader",
    "project manager": "project_leader",
    "project objective": "project_objective",
    "objective": "project_objective",
    "deliverable output": "deliverable_output",
    "deliverable": "deliverable_output",
    "start date": "start_date",
    "start": "start_date",
    "deadline": "deadline",
    "end date": "deadline",
    "completed by": "completed_by_text",
    "project completed by": "completed_by_text",
    "people": "people_count",
    "team size": "people_count",
}


def _xlsx_to_png_bytes(xlsx_bytes: bytes, max_cols: int = 40, max_rows: int = 60) -> bytes:
    """
    Render the first OPPM sheet of an XLSX workbook as a PNG image.

    Uses openpyxl to read cell values and Pillow to draw a compact table.
    Only the first `max_rows` × `max_cols` cells are rendered to keep the
    image small enough for the vision model.
    """
    import openpyxl
    from PIL import Image, ImageDraw, ImageFont

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    # Prefer a sheet whose name contains "oppm"
    target_sheet = wb.active
    for name in wb.sheetnames:
        if "oppm" in name.lower() and "summary" not in name.lower() \
                and "task" not in name.lower() and "member" not in name.lower():
            target_sheet = wb[name]
            break

    rows: list[list[str]] = []
    for row in target_sheet.iter_rows(min_row=1, max_row=max_rows,
                                       min_col=1, max_col=max_cols,
                                       values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])

    if not rows:
        # Return blank white PNG
        img = Image.new("RGB", (400, 100), "white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    # Derive column widths from content
    col_widths = [max(len(rows[r][c]) for r in range(len(rows))) for c in range(len(rows[0]))]
    col_widths = [min(max(w, 4), 20) for w in col_widths]  # clamp 4–20 chars

    FONT_W, FONT_H, PAD = 7, 14, 3
    col_px = [w * FONT_W + PAD * 2 for w in col_widths]
    row_px = FONT_H + PAD * 2

    total_w = sum(col_px)
    total_h = row_px * len(rows)

    img = Image.new("RGB", (total_w, total_h), "white")
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.load_default(size=12)
    except Exception:
        font = ImageFont.load_default()

    # Draw grid lines and cell text
    y = 0
    for r_idx, row in enumerate(rows):
        x = 0
        for c_idx, cell_val in enumerate(row):
            w = col_px[c_idx]
            # Alternate light grey rows for readability
            fill = "#f0f4ff" if r_idx % 2 == 0 else "white"
            draw.rectangle([x, y, x + w - 1, y + row_px - 1], fill=fill, outline="#d0d0d0")
            draw.text((x + PAD, y + PAD), cell_val[:col_widths[c_idx]], fill="black", font=font)
            x += w
        y += row_px

    # Scale up slightly for better OCR quality
    scale = 2
    img = img.resize((total_w * scale, total_h * scale), Image.NEAREST)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=False)
    return buf.getvalue()


# ── PDF → PNG ─────────────────────────────────────────────────────────────────

async def _render_pdf_to_image(file_bytes: bytes) -> bytes:
    """Convert the first page of a PDF to PNG bytes using pdfplumber + Pillow."""
    try:
        import pdfplumber
        from PIL import Image

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                raise ValueError("PDF has no pages")
            page = pdf.pages[0]
            pil_page = page.to_image(resolution=150)
            img: Image.Image = pil_page.original
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return buf.getvalue()
    except ImportError as e:
        raise RuntimeError(f"PDF rendering requires pdfplumber and Pillow: {e}") from e


# ── Main entry point ──────────────────────────────────────────────────────────

async def extract_text_from_upload(
    file_bytes: bytes,
    mime_type: str,
    ollama_url: Optional[str] = None,
) -> OcrResult:
    """
    Stage 1 — send the form to gemma4:31b-cloud for OCR.

    Handles three input types:
      • image (PNG/JPEG/WEBP) → sent directly to vision model
      • PDF                   → rendered to PNG first, then vision model
      • XLSX (Google Sheet)   → rendered to PNG via openpyxl, then vision model

    Args:
        file_bytes: Raw bytes of the file.
        mime_type:  MIME type (image/*, application/pdf, or XLSX mime).
        ollama_url: Override Ollama endpoint (defaults to config).

    Returns:
        OcrResult with raw_text, fields (incl. __form_rules__).
    """
    settings = get_settings()
    base_url = ollama_url or settings.ollama_url

    # ── Determine image bytes to send ─────────────────────────────────────────
    is_xlsx = (
        mime_type in XLSX_MIME_TYPES
        or file_bytes[:4] == XLSX_MAGIC
    )

    if is_xlsx:
        logger.info("OCR: rendering XLSX sheet to PNG for vision model")
        try:
            image_bytes = _xlsx_to_png_bytes(file_bytes)
        except Exception as e:
            raise RuntimeError(f"Failed to render XLSX to image: {e}") from e
    elif mime_type == "application/pdf" or mime_type.endswith("/pdf"):
        logger.info("OCR: converting PDF to image for vision model")
        image_bytes = await _render_pdf_to_image(file_bytes)
    else:
        image_bytes = file_bytes

    image_b64 = base64.b64encode(image_bytes).decode()

    payload = {
        "model": OCR_MODEL_ID,
        "messages": [
            {
                "role": "user",
                "content": _OCR_PROMPT,
                "images": [image_b64],
            }
        ],
        "stream": False,
        "format": "json",
    }

    logger.info("OCR Stage 1: calling %s at %s (source=%s)",
                OCR_MODEL_ID, base_url, "xlsx" if is_xlsx else mime_type)
    try:
        async with httpx.AsyncClient(timeout=240) as client:
            resp = await client.post(f"{base_url}/api/chat", json=payload)
            resp.raise_for_status()
            content = resp.json().get("message", {}).get("content", "")
    except httpx.ConnectError as e:
        raise ProviderUnavailableError("ollama", f"Cannot reach Ollama at {base_url}: {e}") from e
    except httpx.TimeoutException as e:
        raise ProviderUnavailableError("ollama", f"OCR model timed out: {e}") from e
    except httpx.HTTPStatusError as e:
        raise ProviderUnavailableError(
            "ollama", f"HTTP {e.response.status_code}: {e.response.text[:300]}"
        ) from e

    # ── Parse the JSON response ────────────────────────────────────────────────
    raw_data: dict = {}
    try:
        raw_data = json.loads(content)
    except json.JSONDecodeError:
        clean = re.sub(r"^```(?:json)?\s*|\s*```$", "", content, flags=re.DOTALL).strip()
        try:
            raw_data = json.loads(clean)
        except json.JSONDecodeError:
            logger.warning("OCR model returned non-JSON; storing raw as raw_text")
            raw_data = {"raw_text": content, "fields": {}, "form_rules": {}}

    raw_text: str = raw_data.get("raw_text", content)
    fields: dict[str, str] = raw_data.get("fields", {})
    form_rules: dict = raw_data.get("form_rules", {})

    if is_xlsx:
        form_rules.setdefault("notes", "Rendered from XLSX (Google Sheet) via openpyxl")

    fields["__form_rules__"] = json.dumps(form_rules)

    logger.info(
        "OCR Stage 1 complete: %d chars, %d fields detected",
        len(raw_text),
        len(fields) - 1,
    )
    return OcrResult(raw_text=raw_text, fields=fields)
