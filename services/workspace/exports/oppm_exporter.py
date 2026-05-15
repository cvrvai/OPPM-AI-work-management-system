"""OPPM Excel exporter - build_oppm_xlsx(data) -> bytes.

Generates a .xlsx file replicating the classic One Page Project Manager
template layout: title header, project info, task list with sub-objective
checkmarks and timeline dots, gantt/X-diagram section, and bottom
deliverables / forecasts / risks sections.
"""

import io
import logging
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# -- Colour constants ------------------------------------------------
_GREY_BG   = "D9D9D9"
_GREY_GRID = "BFBFBF"
_WHITE     = "FFFFFF"
_NAVY      = "1E3A5F"
_BLUE_700  = "1E40AF"
_GREEN_700 = "166534"
_AMBER_600 = "D97706"
_RED_600   = "DC2626"
_RED_BG    = "FF9999"
_AMBER_BG  = "FFD699"
_LIGHT_GRY = "E6E6E6"

# -- Reusable fills --------------------------------------------------
_F_GREY  = PatternFill("solid", fgColor=_GREY_BG)
_F_WHITE = PatternFill("solid", fgColor=_WHITE)
_F_NAVY  = PatternFill("solid", fgColor=_NAVY)
_F_LGREY = PatternFill("solid", fgColor=_LIGHT_GRY)

# -- Border sides / combos -------------------------------------------
_THIN_BLK = Side(style="thin", color="000000")
_HAIR_GRY = Side(style="hair", color=_GREY_GRID)
_MED_BLK  = Side(style="medium", color="000000")

_B_THIN = Border(left=_THIN_BLK, right=_THIN_BLK, top=_THIN_BLK, bottom=_THIN_BLK)
_B_HAIR = Border(left=_HAIR_GRY, right=_HAIR_GRY, top=_HAIR_GRY, bottom=_HAIR_GRY)
_B_MED  = Border(left=_MED_BLK, right=_MED_BLK, top=_MED_BLK, bottom=_MED_BLK)

# -- Alignments ------------------------------------------------------
_ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_L   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# -- Timeline dot styles ---------------------------------------------
_DOT_MAP: dict[str, tuple[str, Font]] = {
    "planned":     ("\u25a1", Font(name="Calibri", size=10, color="111111")),
    "in_progress": ("\u25cf", Font(name="Calibri", size=10, bold=True, color="111111")),
    "completed":   ("\u25a0", Font(name="Calibri", size=10, bold=True, color="111111")),
    "at_risk":     ("\u25d0", Font(name="Calibri", size=10, bold=True, color=_AMBER_600)),
    "blocked":     ("\u2297", Font(name="Calibri", size=10, bold=True, color=_RED_600)),
}

# -- Owner priority styles -------------------------------------------
_OWNER_STYLES: dict[str, tuple[Font, PatternFill]] = {
    "A": (Font(name="Calibri", size=9, color="111111"), _F_WHITE),
    "B": (Font(name="Calibri", size=9, color="111111"), _F_WHITE),
    "C": (Font(name="Calibri", size=9, color="111111"), _F_WHITE),
}

_OBJ_LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
_FALLBACK_PRIORITY_ORDER = ("A", "B", "C")

# -- Layout constants ------------------------------------------------
NUM_SUBO_COLS  = 6       # sub-objective columns
NUM_TASK_COLS  = 14      # columns merged for task text
MIN_TASK_ROWS  = 16      # pad task section to at least this many rows
GANTT_ROWS     = 36      # rows in the gantt body
BLANK_SEP_ROWS = 4       # blank rows between task list and gantt header
MAX_OWNER_COLS = 6       # classic owner/member columns before the legend zone
LEGEND_COLS    = 4       # fixed right-side legend zone columns
OWNER_BLOCK_WIDTH = 18.0
OWNER_COLUMN_WIDTH = 8.5


# -- Helper functions (mirrors reference script) ---------------------

def _sc(ws, r: int, c: int, val: Any, *,
        font: Font | None = None, align: Alignment | None = None,
        fill: PatternFill | None = None, border: Border | None = None,
        rotation: int | None = None):
    """Set cell value and optional formatting."""
    cell = ws.cell(row=r, column=c)
    cell.value = val
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if rotation is not None:
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=rotation,
        )
    elif align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _fr(ws, r1: int, c1: int, r2: int, c2: int, fill: PatternFill):
    """Fill a rectangular range with a background colour."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill


def _br(ws, r1: int, c1: int, r2: int, c2: int, border: Border):
    """Apply border to every cell in a rectangular range."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


def _mg(ws, r1: int, c1: int, r2: int, c2: int):
    """Merge cells in a rectangular range."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _box(ws, r1: int, c1: int, r2: int, c2: int):
    """Apply a medium outer border and thin inner grid to a box."""
    _br(ws, r1, c1, r2, c2, _B_THIN)
    for c in range(c1, c2 + 1):
        ws.cell(row=r1, column=c).border = Border(
            left=ws.cell(row=r1, column=c).border.left,
            right=ws.cell(row=r1, column=c).border.right,
            top=_MED_BLK,
            bottom=ws.cell(row=r1, column=c).border.bottom,
        )
        ws.cell(row=r2, column=c).border = Border(
            left=ws.cell(row=r2, column=c).border.left,
            right=ws.cell(row=r2, column=c).border.right,
            top=ws.cell(row=r2, column=c).border.top,
            bottom=_MED_BLK,
        )
    for r in range(r1, r2 + 1):
        ws.cell(row=r, column=c1).border = Border(
            left=_MED_BLK,
            right=ws.cell(row=r, column=c1).border.right,
            top=ws.cell(row=r, column=c1).border.top,
            bottom=ws.cell(row=r, column=c1).border.bottom,
        )
        ws.cell(row=r, column=c2).border = Border(
            left=ws.cell(row=r, column=c2).border.left,
            right=_MED_BLK,
            top=ws.cell(row=r, column=c2).border.top,
            bottom=ws.cell(row=r, column=c2).border.bottom,
        )


def _draw_right_legend_block(
    ws,
    *,
    title: str,
    rows: list[tuple[str, str]],
    start_row: int,
    start_col: int,
    end_col: int,
    title_font: Font,
    body_font: Font,
):
    """Draw a fixed right-side legend block matching the preview layout."""
    body_start_col = min(end_col, start_col + 1)

    _mg(ws, start_row, start_col, start_row, end_col)
    _sc(ws, start_row, start_col, title, font=title_font, align=_ALIGN_C, fill=_F_WHITE)

    for offset, (symbol, label) in enumerate(rows, start=1):
        row = start_row + offset
        _sc(ws, row, start_col, symbol, font=body_font, align=_ALIGN_C, fill=_F_WHITE)
        if body_start_col <= end_col:
            _mg(ws, row, body_start_col, row, end_col)
            _sc(ws, row, body_start_col, label, font=body_font, align=_ALIGN_L, fill=_F_WHITE)

    _box(ws, start_row, start_col, start_row + len(rows), end_col)


def _fmt_date(val) -> str:
    """Format a date value to M/D/YY string."""
    if not val:
        return "\u2014"
    if isinstance(val, date):
        return f"{val.month}/{val.day}/{str(val.year)[-2:]}"
    if isinstance(val, str):
        try:
            d = date.fromisoformat(val[:10])
            return f"{d.month}/{d.day}/{str(d.year)[-2:]}"
        except (ValueError, IndexError):
            return val
    return str(val)


# -- Main builder ----------------------------------------------------

def build_oppm_xlsx(data: dict) -> bytes:
    """Build a .xlsx workbook replicating the classic OPPM one-page layout.

    Parameters
    ----------
    data : dict
        Expected keys: project, objectives, sub_objectives, members,
        timeline, weeks, costs, deliverables, forecasts, risks.

    Returns
    -------
    bytes
        The .xlsx file content.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "OPPM"

    # -- Extract data -------------------------------------------------
    project      = data.get("project", {})
    header       = data.get("header") or {}
    objectives   = data.get("objectives", [])
    sub_objs     = data.get("sub_objectives", [])
    members      = (data.get("project_members") or data.get("members") or [])[:MAX_OWNER_COLS]
    timeline     = data.get("timeline", [])
    weeks        = data.get("weeks", [])
    costs        = data.get("costs", {})
    deliverables = data.get("deliverables", [])
    forecasts    = data.get("forecasts", [])
    risks        = data.get("risks", [])

    num_weeks   = len(weeks)
    num_members = len(members)

    fallback_owner_by_user_id: dict[str, dict[str, str]] = {}
    for index, member in enumerate(members):
        user_id = member.get("user_id")
        member_id = member.get("id")
        if not user_id or not member_id:
            continue
        fallback_owner_by_user_id[str(user_id)] = {
            "member_id": str(member_id),
            "priority": _FALLBACK_PRIORITY_ORDER[index] if index < len(_FALLBACK_PRIORITY_ORDER) else "A",
        }

    # -- Sub-objective position map -----------------------------------
    sub_obj_by_pos: dict[int, dict] = {}
    for so in sub_objs:
        pos = so.get("position", 0)
        if 1 <= pos <= 6:
            sub_obj_by_pos[pos] = so

    # -- Timeline lookup: task_id -> week_start -> entry --------------
    tl_map: dict[str, dict[str, dict]] = {}
    for entry in timeline:
        tid = str(entry.get("task_id", ""))
        ws_date = str(entry.get("week_start", ""))
        if tid and ws_date:
            tl_map.setdefault(tid, {})[ws_date] = {
                "status": entry.get("status", ""),
                "quality": entry.get("quality"),
            }

    # -- Column layout ------------------------------------------------
    C_SUBO_S = 1
    C_SUBO_E = NUM_SUBO_COLS                           # 6
    C_TASK_S = C_SUBO_E + 1                             # 7
    C_TASK_E = C_TASK_S + NUM_TASK_COLS - 1             # 20

    date_cols = max(num_weeks, 1)
    mem_cols  = max(num_members, 1)
    visible_owner_cols = max(num_members, 1)

    C_DATE_S = C_TASK_E + 1                             # 21
    C_DATE_E = C_DATE_S + date_cols - 1
    C_MEM_S        = C_DATE_E + 1
    C_MEM_E        = C_MEM_S + MAX_OWNER_COLS - 1
    C_LEGEND_S     = C_MEM_E + 1
    C_LEGEND_E     = C_LEGEND_S + LEGEND_COLS - 1
    LAST_COL       = C_LEGEND_E

    # -- Column widths ------------------------------------------------
    for c in range(C_SUBO_S, C_SUBO_E + 1):
        ws.column_dimensions[get_column_letter(c)].width = 3.2
    for c in range(C_TASK_S, C_TASK_E + 1):
        ws.column_dimensions[get_column_letter(c)].width = 3.5
    ws.column_dimensions[get_column_letter(C_TASK_S)].width = 8
    for c in range(C_DATE_S, C_DATE_E + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5
    owner_column_width = OWNER_COLUMN_WIDTH
    for i, c in enumerate(range(C_MEM_S, C_MEM_E + 1)):
        dim = ws.column_dimensions[get_column_letter(c)]
        dim.width = owner_column_width
        dim.hidden = i >= visible_owner_cols
    legend_widths = [5.5, 8.5, 8.5, 8.5]
    for offset, c in enumerate(range(C_LEGEND_S, C_LEGEND_E + 1)):
        ws.column_dimensions[get_column_letter(c)].width = legend_widths[offset] if offset < len(legend_widths) else 8.5

    # -- Font shortcuts -----------------------------------------------
    f_title      = Font(name="Calibri", size=14, bold=True)
    f_bold       = Font(name="Calibri", size=10, bold=True)
    f_normal     = Font(name="Calibri", size=10)
    f_header     = Font(name="Calibri", size=9, bold=True)
    f_obj_hdr    = Font(name="Calibri", size=10, bold=True, color=_WHITE)
    f_main_task  = Font(name="Calibri", size=9, bold=True)
    f_task       = Font(name="Calibri", size=9)
    f_check      = Font(name="Calibri", size=10, bold=True, color=_GREEN_700)
    f_rotated    = Font(name="Calibri", size=8)
    f_x_label    = Font(name="Calibri", size=11, bold=True, color="444444")
    f_section    = Font(name="Calibri", size=10, bold=True)
    f_risk_white = Font(name="Calibri", size=9, bold=True, color=_WHITE)
    f_subo_label = Font(name="Calibri", size=8, bold=True)

    # =================================================================
    # ROWS 1-2: Title header
    # =================================================================
    lead_name = header.get("project_leader_text") or project.get("lead_name") or "\u2014"
    proj_title = project.get("title") or "\u2014"

    _mg(ws, 1, C_SUBO_S, 2, C_TASK_E)
    _sc(ws, 1, C_SUBO_S,
        f"Project Leader: {lead_name}",
        font=f_title, fill=_F_GREY, align=_ALIGN_L)
    _fr(ws, 1, C_SUBO_S, 2, C_TASK_E, _F_GREY)

    _mg(ws, 1, C_DATE_S, 2, LAST_COL)
    _sc(ws, 1, C_DATE_S,
        f"Project Name: {proj_title}",
        font=f_title, fill=_F_GREY, align=_ALIGN_L)
    _fr(ws, 1, C_DATE_S, 2, LAST_COL, _F_GREY)

    _br(ws, 1, C_SUBO_S, 2, LAST_COL, _B_MED)

    # =================================================================
    # ROWS 3-6: Project info
    # =================================================================
    info_rows = [
        ("Project Objective:",
         project.get("objective_summary") or project.get("description") or "\u2014"),
        ("Deliverable Output:",
         ", ".join(d.get("description", "") for d in deliverables if d.get("description")) or "\u2014"),
        ("Start Date:", _fmt_date(project.get("start_date"))),
        ("Deadline:",   _fmt_date(project.get("deadline"))),
    ]
    for i, (label, value) in enumerate(info_rows):
        r = 3 + i
        _mg(ws, r, C_SUBO_S, r, C_SUBO_E)
        _sc(ws, r, C_SUBO_S, label, font=f_bold, align=_ALIGN_L, border=_B_THIN)
        _fr(ws, r, C_SUBO_S, r, C_SUBO_E, _F_WHITE)
        _mg(ws, r, C_TASK_S, r, LAST_COL)
        _sc(ws, r, C_TASK_S, value, font=f_normal, align=_ALIGN_L, border=_B_THIN)

    # Medium border around info block
    for r in range(3, 7):
        ws.cell(row=r, column=C_SUBO_S).border = Border(
            left=_MED_BLK, top=_THIN_BLK, bottom=_THIN_BLK, right=_THIN_BLK)
        ws.cell(row=r, column=LAST_COL).border = Border(
            right=_MED_BLK, top=_THIN_BLK, bottom=_THIN_BLK, left=_THIN_BLK)
    _br(ws, 3, C_SUBO_S, 3, LAST_COL, Border(
        left=_MED_BLK, right=_MED_BLK, top=_MED_BLK, bottom=_THIN_BLK))
    _br(ws, 6, C_SUBO_S, 6, LAST_COL, Border(
        left=_MED_BLK, right=_MED_BLK, top=_THIN_BLK, bottom=_MED_BLK))

    # =================================================================
    # ROW 7: Column headers
    # =================================================================
    r = 7

    _mg(ws, r, C_SUBO_S, r, C_SUBO_E)
    _sc(ws, r, C_SUBO_S, "Sub Objective",
        font=f_header, fill=_F_GREY, align=_ALIGN_C, border=_B_THIN)
    _fr(ws, r, C_SUBO_S, r, C_SUBO_E, _F_GREY)

    _mg(ws, r, C_TASK_S, r, C_TASK_E)
    _sc(ws, r, C_TASK_S, "Major Tasks (Deadline)",
        font=f_header, fill=_F_GREY, align=_ALIGN_C, border=_B_THIN)
    _fr(ws, r, C_TASK_S, r, C_TASK_E, _F_GREY)

    completed_by_text = header.get("completed_by_text") or _fmt_date(project.get("deadline"))
    date_header = f"Project Completed By: {completed_by_text}"
    _mg(ws, r, C_DATE_S, r, C_DATE_E)
    _sc(ws, r, C_DATE_S, date_header,
        font=f_header, fill=_F_GREY, align=_ALIGN_C, border=_B_THIN)
    _fr(ws, r, C_DATE_S, r, C_DATE_E, _F_GREY)

    _mg(ws, r, C_MEM_S, r, C_MEM_E)
    _sc(ws, r, C_MEM_S, "Owner / Priority",
        font=f_header, fill=_F_GREY, align=_ALIGN_C, border=_B_THIN)
    _fr(ws, r, C_MEM_S, r, C_MEM_E, _F_GREY)

    for c in range(C_LEGEND_S, C_LEGEND_E + 1):
        _sc(ws, r, c, None, fill=_F_GREY, border=_B_THIN)

    # =================================================================
    # ROWS 8+: Task rows (objectives + tasks)
    # =================================================================
    task_start_row = 8
    r = task_start_row

    for obj_idx, obj in enumerate(objectives):
        tasks = obj.get("tasks", [])

        # -- Objective header row (navy bar) --------------------------
        for c in range(C_SUBO_S, C_SUBO_E + 1):
            _sc(ws, r, c, None, fill=_F_NAVY, border=_B_THIN)

        obj_title = obj.get("title") or "\u2014"
        _mg(ws, r, C_TASK_S, r, C_TASK_E)
        _sc(ws, r, C_TASK_S,
            f"{obj_idx + 1}. {obj_title}",
            font=f_obj_hdr, fill=_F_NAVY, align=_ALIGN_L, border=_B_THIN)
        _fr(ws, r, C_TASK_S, r, C_TASK_E, _F_NAVY)

        for c in range(C_DATE_S, C_DATE_E + 1):
            _sc(ws, r, c, None, fill=_F_NAVY, border=_B_THIN)
        for c in range(C_MEM_S, C_MEM_E + 1):
            _sc(ws, r, c, None, fill=_F_NAVY, border=_B_THIN)
        r += 1

        # -- Separate main tasks and sub-tasks for hierarchical numbering --
        main_tasks = [t for t in tasks if not t.get("parent_task_id")]
        sub_tasks_by_parent: dict[str, list[dict]] = {}
        for t in tasks:
            pid = t.get("parent_task_id")
            if pid:
                sub_tasks_by_parent.setdefault(str(pid), []).append(t)

        # Build ordered list: main task followed by its sub-tasks
        ordered_tasks: list[tuple[dict, bool, str]] = []  # (task, is_sub, label)
        main_idx = 0
        for mt in main_tasks:
            main_idx += 1
            label = f"{obj_idx + 1}.{main_idx}"
            ordered_tasks.append((mt, False, label))
            children = sub_tasks_by_parent.get(str(mt.get("id", "")), [])
            for sub_idx, st in enumerate(children, 1):
                sub_label = f"{obj_idx + 1}.{main_idx}.{sub_idx}"
                ordered_tasks.append((st, True, sub_label))

        # If no main tasks found (all tasks are flat), fall back to flat numbering
        if not main_tasks:
            for flat_idx, t in enumerate(tasks, 1):
                ordered_tasks.append((t, False, f"{obj_idx + 1}.{flat_idx}"))

        # -- Task rows ------------------------------------------------
        for task, is_sub, label in ordered_tasks:
            task_id = str(task.get("id", ""))
            task_sub_ids = {str(s) for s in task.get("sub_objective_ids", [])}
            task_owners: dict[str, str] = {}
            for own in task.get("owners", []):
                task_owners[str(own.get("member_id", ""))] = own.get("priority", "")
            if not task_owners and task.get("assignee_id"):
                fallback_owner = fallback_owner_by_user_id.get(str(task.get("assignee_id")))
                if fallback_owner:
                    task_owners[fallback_owner["member_id"]] = fallback_owner["priority"]

            row_fill = _F_WHITE if is_sub else _F_LGREY
            row_font = f_task if is_sub else f_main_task

            # Sub-objective checkmarks
            for i in range(NUM_SUBO_COLS):
                pos = i + 1
                so = sub_obj_by_pos.get(pos)
                if so and so["id"] in task_sub_ids:
                    _sc(ws, r, C_SUBO_S + i, "\u2713",
                        font=f_check, fill=row_fill, align=_ALIGN_C, border=_B_HAIR)
                else:
                    _sc(ws, r, C_SUBO_S + i, None,
                        fill=row_fill, border=_B_HAIR)

            # Task text (merged) with inline deadline
            due = task.get("due_date")
            task_text = task.get("title", "")
            if due:
                task_text += f"  ({_fmt_date(due)})"
            indent = "      " if is_sub else "   "
            _mg(ws, r, C_TASK_S, r, C_TASK_E)
            _sc(ws, r, C_TASK_S,
                f"{indent}{label}  {task_text}",
                font=row_font, fill=row_fill, align=_ALIGN_L, border=_B_HAIR)

            # Timeline dots / project identity symbols
            task_tl = tl_map.get(task_id, {})
            fallback_status = str(task.get("status", "planned") or "planned")
            if fallback_status == "todo":
                fallback_status = "planned"
            fallback_dot = _DOT_MAP.get(fallback_status, _DOT_MAP["planned"])
            fallback_col: int | None = None
            if not task_tl and task.get("due_date"):
                try:
                    due_date = date.fromisoformat(str(task.get("due_date"))[:10])
                    best_diff = None
                    for i, week in enumerate(weeks):
                        week_start = date.fromisoformat(str(week.get("start", ""))[:10])
                        diff = abs((week_start - due_date).days)
                        if best_diff is None or diff < best_diff:
                            best_diff = diff
                            fallback_col = C_DATE_S + i
                except (TypeError, ValueError):
                    fallback_col = None
            for i, week in enumerate(weeks):
                col = C_DATE_S + i
                ws_date = week.get("start", "")
                entry = task_tl.get(ws_date, {})
                status = entry.get("status", "")
                dot_info = _DOT_MAP.get(status)
                if dot_info:
                    symbol, dot_font = dot_info
                    _sc(ws, r, col, symbol,
                        font=dot_font, fill=row_fill, align=_ALIGN_C, border=_B_HAIR)
                elif fallback_col == col:
                    if fallback_col is not None:
                        symbol, dot_font = fallback_dot
                        _sc(ws, r, col, symbol,
                            font=dot_font, fill=row_fill, align=_ALIGN_C, border=_B_HAIR)
                else:
                    _sc(ws, r, col, None, fill=row_fill, border=_B_HAIR)
            for c in range(C_DATE_S + num_weeks, C_DATE_E + 1):
                _sc(ws, r, c, None, fill=row_fill, border=_B_HAIR)

            # Member owner columns (A / B / C)
            for i, member in enumerate(members):
                col = C_MEM_S + i
                mid = str(member.get("id", ""))
                priority = task_owners.get(mid, "")
                style = _OWNER_STYLES.get(priority)
                if style:
                    _sc(ws, r, col, priority,
                        font=style[0], fill=row_fill, align=_ALIGN_C, border=_B_HAIR)
                else:
                    _sc(ws, r, col, None, fill=row_fill, border=_B_HAIR)
            for c in range(C_MEM_S + visible_owner_cols, C_MEM_E + 1):
                _sc(ws, r, c, None, fill=row_fill, border=_B_HAIR)
            for c in range(C_LEGEND_S, C_LEGEND_E + 1):
                _sc(ws, r, c, None, fill=row_fill, border=_B_HAIR)

            r += 1

    # Pad to minimum task rows
    task_end_row = task_start_row + MIN_TASK_ROWS - 1
    while r <= task_end_row:
        for c in range(C_SUBO_S, LAST_COL + 1):
            _sc(ws, r, c, None, fill=_F_WHITE, border=_B_HAIR)
        r += 1

    # Thin border around entire task block
    _br(ws, task_start_row, C_SUBO_S, r - 1, LAST_COL, _B_THIN)

    _draw_right_legend_block(
        ws,
        title="Priority",
        rows=[
            ("A", "Primary"),
            ("B", "Secondary"),
            ("C", "Support"),
        ],
        start_row=task_start_row + 1,
        start_col=C_LEGEND_S,
        end_col=C_LEGEND_E,
        title_font=f_header,
        body_font=f_task,
    )

    # =================================================================
    # BLANK SEPARATOR ROWS
    # =================================================================
    separator_start_row = r
    separator_end_row = r + BLANK_SEP_ROWS - 1
    for sep_row in range(separator_start_row, separator_end_row + 1):
        for c in range(C_SUBO_S, LAST_COL + 1):
            _sc(ws, sep_row, c, None, fill=_F_WHITE, border=_B_HAIR)

    _draw_right_legend_block(
        ws,
        title="Project Identity Symbol",
        rows=[
            ("□", "Start"),
            ("●", "In Progress"),
            ("■", "Complete"),
        ],
        start_row=separator_start_row,
        start_col=C_LEGEND_S,
        end_col=C_LEGEND_E,
        title_font=f_header,
        body_font=f_task,
    )

    r = separator_end_row + 1

    # =================================================================
    # "# People working on the project"
    # =================================================================
    people_count = header.get("people_count")
    if people_count is None:
        people_count = len([member for member in members if member.get("id") or member.get("display_name")])
    people_label = f"# People working on the project: {people_count}" if people_count else "# People working on the project"
    _mg(ws, r, C_SUBO_S, r, LAST_COL)
    _sc(ws, r, C_SUBO_S, people_label,
        font=f_bold, fill=_F_GREY, align=_ALIGN_L, border=_B_THIN)
    _fr(ws, r, C_SUBO_S, r, LAST_COL, _F_GREY)
    r += 1

    # =================================================================
    # GANTT HEADER ROW (rotated text)
    # =================================================================
    gantt_hdr_row = r
    ws.row_dimensions[r].height = 60

    for i in range(NUM_SUBO_COLS):
        _sc(ws, r, C_SUBO_S + i, str(i + 1),
            font=f_rotated, fill=_F_GREY, align=_ALIGN_C, border=_B_THIN,
            rotation=90)

    _mg(ws, r, C_TASK_S, r, C_TASK_E)
    _sc(ws, r, C_TASK_S, None, fill=_F_GREY, border=_B_THIN)
    _fr(ws, r, C_TASK_S, r, C_TASK_E, _F_GREY)

    for i, week in enumerate(weeks):
        _sc(ws, r, C_DATE_S + i, week.get("label", ""),
            font=f_rotated, fill=_F_GREY, border=_B_THIN, rotation=90)
    for c in range(C_DATE_S + num_weeks, C_DATE_E + 1):
        _sc(ws, r, c, None, fill=_F_GREY, border=_B_THIN)

    for i, member in enumerate(members):
        name = (
            member.get("display_name") or
            member.get("name") or
            member.get("email", "").split("@")[0] or
            f"Member {i + 1}"
        )[:12]
        _sc(ws, r, C_MEM_S + i, name,
            font=f_rotated, fill=_F_GREY, border=_B_THIN, rotation=90)
    for c in range(C_MEM_S + visible_owner_cols, C_MEM_E + 1):
        _sc(ws, r, c, None, fill=_F_GREY, border=_B_THIN)
    for c in range(C_LEGEND_S, C_LEGEND_E + 1):
        _sc(ws, r, c, None, fill=_F_GREY, border=_B_THIN)

    r += 1

    # =================================================================
    # GANTT BODY: sub-obj labels + X-diagram + grid
    # =================================================================
    gantt_body_start = r
    gantt_body_end   = r + GANTT_ROWS - 1

    _fr(ws, gantt_body_start, C_SUBO_S, gantt_body_end, LAST_COL, _F_WHITE)
    _br(ws, gantt_body_start, C_SUBO_S, gantt_body_end, LAST_COL, _B_HAIR)

    # -- Sub-objective labels (each column full height, rotated) ------
    for i in range(NUM_SUBO_COLS):
        pos = i + 1
        so = sub_obj_by_pos.get(pos)
        label = so.get("title", f"Sub Obj {pos}") if so else f"Sub Obj {pos}"
        _mg(ws, gantt_body_start, C_SUBO_S + i,
            gantt_body_end, C_SUBO_S + i)
        _sc(ws, gantt_body_start, C_SUBO_S + i, label,
            font=f_subo_label, fill=_F_LGREY, border=_B_THIN, rotation=90)

    # -- X-diagram labels in center area ------------------------------
    mid_r = gantt_body_start + GANTT_ROWS // 2
    mid_c = (C_TASK_S + C_TASK_E) // 2

    x_labels = [
        ("Major Tasks",
         gantt_body_start + 2, C_TASK_S,
         gantt_body_start + 4, mid_c - 1),
        ("Target Dates",
         gantt_body_start + 2, mid_c + 1,
         gantt_body_start + 4, C_TASK_E),
        ("Summary &\nForecast",
         mid_r - 1, mid_c - 3,
         mid_r + 1, mid_c + 3),
        ("Sub Objectives",
         gantt_body_end - 4, C_TASK_S,
         gantt_body_end - 2, mid_c - 1),
        ("Costs",
         gantt_body_end - 4, mid_c + 1,
         gantt_body_end - 2, C_TASK_E),
    ]
    for label, r1, c1, r2, c2 in x_labels:
        _mg(ws, r1, c1, r2, c2)
        _sc(ws, r1, c1, label,
            font=f_x_label, fill=_F_WHITE,
            align=Alignment(horizontal="center", vertical="center",
                            wrap_text=True))

    _br(ws, gantt_body_start, C_DATE_S, gantt_body_end, C_DATE_E, _B_HAIR)
    _br(ws, gantt_body_start, C_MEM_S,  gantt_body_end, C_MEM_E,  _B_HAIR)
    _br(ws, gantt_body_start, C_LEGEND_S, gantt_body_end, C_LEGEND_E, _B_HAIR)

    _br(ws, gantt_hdr_row, C_SUBO_S, gantt_body_end, LAST_COL, _B_THIN)

    r = gantt_body_end + 2

    # =================================================================
    # BOTTOM SECTION: Summary Deliverable / Forecast / Risk
    # =================================================================

    # -- Summary Deliverables -----------------------------------------
    _mg(ws, r, C_SUBO_S, r, LAST_COL)
    _sc(ws, r, C_SUBO_S, "Summary Deliverable",
        font=f_section, fill=_F_GREY, align=_ALIGN_L, border=_B_THIN)
    _fr(ws, r, C_SUBO_S, r, LAST_COL, _F_GREY)
    r += 1

    for d in deliverables:
        _mg(ws, r, C_SUBO_S, r, LAST_COL)
        _sc(ws, r, C_SUBO_S,
            f"  {d.get('item_number', '')}. {d.get('description', '')}",
            font=f_normal, fill=_F_WHITE, align=_ALIGN_L, border=_B_HAIR)
        r += 1

    # -- Forecast -----------------------------------------------------
    _mg(ws, r, C_SUBO_S, r, LAST_COL)
    _sc(ws, r, C_SUBO_S, "Forecast",
        font=f_section, fill=_F_GREY, align=_ALIGN_L, border=_B_THIN)
    _fr(ws, r, C_SUBO_S, r, LAST_COL, _F_GREY)
    r += 1

    for f in forecasts:
        _mg(ws, r, C_SUBO_S, r, LAST_COL)
        _sc(ws, r, C_SUBO_S,
            f"  {f.get('item_number', '')}. {f.get('description', '')}",
            font=f_normal, fill=_F_WHITE, align=_ALIGN_L, border=_B_HAIR)
        r += 1

    # -- Risks --------------------------------------------------------
    _mg(ws, r, C_SUBO_S, r, LAST_COL)
    _sc(ws, r, C_SUBO_S, "Risk",
        font=f_section, fill=_F_GREY, align=_ALIGN_L, border=_B_THIN)
    _fr(ws, r, C_SUBO_S, r, LAST_COL, _F_GREY)
    r += 1

    for risk in risks:
        rag = risk.get("rag", "green")
        if rag == "red":
            rag_fill = PatternFill("solid", fgColor=_RED_BG)
            rag_font = f_risk_white
        elif rag == "amber":
            rag_fill = PatternFill("solid", fgColor=_AMBER_BG)
            rag_font = f_risk_white
        else:
            rag_fill = _F_WHITE
            rag_font = f_normal

        _mg(ws, r, C_SUBO_S, r, LAST_COL)
        _sc(ws, r, C_SUBO_S,
            f"  [{rag.upper()}] {risk.get('description', '')}",
            font=rag_font, fill=rag_fill, align=_ALIGN_L, border=_B_HAIR)
        _fr(ws, r, C_SUBO_S, r, LAST_COL, rag_fill)
        r += 1

    # =================================================================
    # Legend rows
    # =================================================================
    r += 1
    _sc(ws, r, C_SUBO_S, "Legend:",
        font=Font(name="Calibri", size=9, bold=True, color="999999"))
    legend = [
        ("\u25a1 Planned", _DOT_MAP["planned"][1]),
        ("\u25cf In Progress", _DOT_MAP["in_progress"][1]),
        ("\u25a0 Completed", _DOT_MAP["completed"][1]),
        ("\u25d0 At Risk", _DOT_MAP["at_risk"][1]),
        ("\u2297 Blocked", _DOT_MAP["blocked"][1]),
    ]
    for li, (text, fnt) in enumerate(legend):
        _sc(ws, r, C_SUBO_S + 1 + li, text, font=fnt)

    r += 1
    _sc(ws, r, C_SUBO_S, "Owners:",
        font=Font(name="Calibri", size=9, bold=True, color="999999"))
    for priority, (fnt, fll) in _OWNER_STYLES.items():
        label_map = {"A": "Primary", "B": "Secondary", "C": "Support"}
        col_offset = {"A": 1, "B": 2, "C": 3}[priority]
        _sc(ws, r, C_SUBO_S + col_offset,
            f"{priority} = {label_map[priority]}", font=fnt, fill=fll)

    # =================================================================
    # Page / freeze setup
    # =================================================================
    ws.freeze_panes = ws.cell(row=8, column=C_TASK_S)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── OPPM Import Template ──────────────────────────────────────────────────────

_GREY_TEXT = "9CA3AF"


def _tmpl_header(ws: Any, row: int, col: int, value: str, width: float | None = None) -> Any:
    """Write a navy bold header cell to an import-template sheet."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=9, bold=True, color=_WHITE)
    cell.fill = _F_NAVY
    cell.alignment = _ALIGN_C
    if width is not None:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _tmpl_sample(ws: Any, row: int, col: int, value: str, center: bool = False) -> Any:
    """Write a grey italic sample-data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=9, italic=True, color=_GREY_TEXT)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    return cell


def build_oppm_template(project_title: str = "") -> bytes:
    """Build a blank OPPM import template .xlsx for users to fill in and upload.

    Sheets:
      Instructions   - usage guide
      Tasks          - Objective | Task Name | Due Date | Sub Obj 1-6
      Sub Objectives - Position (1-6) | Label
      Deliverables   - # | Description
      Forecasts      - # | Description
      Risks          - # | Description | RAG (green/amber/red)
    """
    wb = Workbook()

    # ── Sheet 1: Instructions ─────────────────────────────────
    ws_i = wb.active
    ws_i.title = "Instructions"
    ws_i.merge_cells("A1:E1")
    c = ws_i["A1"]
    c.value = "OPPM Import Template — Instructions"
    c.font = Font(name="Calibri", size=13, bold=True, color=_WHITE)
    c.fill = _F_NAVY
    c.alignment = _ALIGN_C
    ws_i.row_dimensions[1].height = 28
    ws_i.column_dimensions["A"].width = 90

    for ri, (txt, bold) in enumerate([
        ("", False),
        ("How to use this template:", True),
        ("", False),
        ("1. Fill in the 'Tasks' sheet with one task per row.", False),
        ("   Column A (Objective)        Enter the objective name once per group. Leave blank for follow-on tasks.", False),
        ("   Column B (Task Name)        Required.", False),
        ("   Column C (Due Date)         Optional. Format: YYYY-MM-DD  e.g. 2026-05-01", False),
        ("   Columns D-I (Sub Obj 1-6)   Put x to link the task to that sub-objective.", False),
        ("", False),
        ("2. Optionally fill the 'Sub Objectives' sheet to define position labels (1-6).", False),
        ("3. Fill 'Deliverables', 'Forecasts', and 'Risks' sheets as needed.", False),
        ("   For Risks, Column C accepts: green / amber / red  (default: green).", False),
        ("4. Save this file, then click  Import XLSX  on the OPPM page and select it.", False),
        ("", False),
        ("Notes:", True),
        ("* Import is APPEND-ONLY. Existing data is never deleted.", False),
        ("* Objectives matched by name (case-insensitive) will be reused, not duplicated.", False),
    ], start=2):
        cx = ws_i.cell(row=ri, column=1, value=txt)
        cx.font = Font(name="Calibri", size=10, bold=bold, color="111827")

    # ── Sheet 2: Tasks ────────────────────────────────────────
    ws_t = wb.create_sheet("Tasks")
    ws_t.row_dimensions[1].height = 30
    for ci, (h, w) in enumerate([
        ("Objective", 28), ("Task Name", 34), ("Due Date (YYYY-MM-DD)", 20),
        ("Sub Obj 1", 11), ("Sub Obj 2", 11), ("Sub Obj 3", 11),
        ("Sub Obj 4", 11), ("Sub Obj 5", 11), ("Sub Obj 6", 11),
    ], start=1):
        _tmpl_header(ws_t, 1, ci, h, w)

    for ri, row_data in enumerate([
        ("Design & UX",  "Homepage wireframe",       "2026-04-15", "x", "",  "",  "",  "",  ""),
        ("",             "Navigation component",      "2026-04-20", "x", "x", "",  "",  "",  ""),
        ("",             "Mobile-responsive layout",  "2026-04-25", "",  "x", "",  "",  "",  ""),
        ("Backend API",  "Auth endpoints",            "2026-05-01", "",  "",  "x", "",  "",  ""),
        ("",             "Database schema",           "2026-05-05", "",  "",  "x", "",  "",  ""),
        ("Testing",      "Unit tests",                "2026-05-10", "",  "",  "",  "x", "",  ""),
    ], start=2):
        for ci, val in enumerate(row_data, start=1):
            _tmpl_sample(ws_t, ri, ci, str(val), center=(ci > 2))

    note = ws_t.cell(row=10, column=1, value="← Delete sample rows; your data starts at row 2")
    note.font = Font(name="Calibri", size=8, italic=True, color=_GREY_TEXT)

    # ── Sheet 3: Sub Objectives ───────────────────────────────
    ws_so = wb.create_sheet("Sub Objectives")
    ws_so.row_dimensions[1].height = 22
    _tmpl_header(ws_so, 1, 1, "Position (1-6)", 16)
    _tmpl_header(ws_so, 1, 2, "Label", 34)
    for ri, (pos, lbl) in enumerate(
        [(1, "Quality"), (2, "Schedule"), (3, "Cost"), (4, "Scope"), (5, "Risk"), (6, "Team")],
        start=2,
    ):
        _tmpl_sample(ws_so, ri, 1, str(pos), center=True)
        _tmpl_sample(ws_so, ri, 2, lbl)

    # ── Sheets 4-6: Deliverables / Forecasts / Risks ─────────
    for sheet_name, extra in (
        ("Deliverables", None),
        ("Forecasts", None),
        ("Risks", "RAG (green/amber/red)"),
    ):
        ws_x = wb.create_sheet(sheet_name)
        ws_x.row_dimensions[1].height = 22
        _tmpl_header(ws_x, 1, 1, "#", 6)
        _tmpl_header(ws_x, 1, 2, "Description", 52)
        if extra:
            _tmpl_header(ws_x, 1, 3, extra, 24)
        row_vals: list[str] = ["1", f"Sample {sheet_name.lower()} item"] + (["green"] if extra else [])
        for ci, val in enumerate(row_vals, start=1):
            _tmpl_sample(ws_x, 2, ci, val, center=(ci != 2))

    buf2 = io.BytesIO()
    wb.save(buf2)
    logger.info("OPPM template built for project %r", project_title)
    return buf2.getvalue()
