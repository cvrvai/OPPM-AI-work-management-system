"""Export service — delegates to oppm_service for data assembly and calls the OPPM exporter."""

import io
import logging
import uuid as _uuid
from datetime import date as _date

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from exports.oppm_exporter import build_oppm_xlsx
from domains.oppm.repository import (
    DeliverableRepository, ForecastRepository, ObjectiveRepository,
    RiskRepository, SubObjectiveRepository,
)
from domains.project.repository import ProjectRepository
from domains.oppm.service import get_oppm_data
from shared.models.oppm import TaskSubObjective
from shared.models.task import Task

logger = logging.getLogger(__name__)

_TRUTHY = {"x", "1", "y", "yes", "true", "v", "check", "●", "•"}
_VALID_RAG = {"green", "amber", "red"}


async def export_oppm_xlsx(session: AsyncSession, project_id: str, workspace_id: str) -> bytes:
    """Assemble OPPM data via the shared service function and return .xlsx bytes."""
    data = await get_oppm_data(session, project_id, workspace_id)
    return build_oppm_xlsx(data)


def parse_oppm_xlsx_to_preview(xlsx_bytes: bytes) -> dict:
    """Parse an OPPM import template .xlsx and return structured JSON preview.

    This is a *pure* function — no session, no DB writes.  The returned dict
    has the same shape as the AI OCR extractor response so the same review
    drawer can be reused in the frontend.

    Raises HTTPException(400) on corrupt/invalid file.
    """

    def _truthy(v: object) -> bool:
        return str(v).strip().lower() in _TRUTHY if v else False

    def _pdate(v: object) -> str | None:
        if v is None:
            return None
        if isinstance(v, _date):
            return v.isoformat()
        try:
            return _date.fromisoformat(str(v).strip()[:10]).isoformat()
        except ValueError:
            return None

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt XLSX file")

    preview: dict = {
        "sub_objectives": [],
        "objectives": [],
        "deliverables": [],
        "forecasts": [],
        "risks": [],
    }

    # ── Sub Objectives ────────────────────────────────────────
    if "Sub Objectives" in wb.sheetnames:
        seen_pos: set[int] = set()
        for row in wb["Sub Objectives"].iter_rows(min_row=2, values_only=True):
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                break
            try:
                pos = int(row[0])
            except (TypeError, ValueError):
                continue
            label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if 1 <= pos <= 6 and label and pos not in seen_pos:
                preview["sub_objectives"].append({"position": pos, "label": label})
                seen_pos.add(pos)

    # ── Tasks (objectives + tasks) ────────────────────────────
    if "Tasks" in wb.sheetnames:
        cur_obj: dict | None = None
        for row in wb["Tasks"].iter_rows(min_row=2, values_only=True):
            row_vals = row[:9] if len(row) >= 9 else row
            if all(v is None or str(v).strip() == "" for v in row_vals):
                break

            obj_raw  = str(row[0]).strip() if row[0] else ""
            task_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            if obj_raw:
                cur_obj = {"title": obj_raw, "tasks": []}
                preview["objectives"].append(cur_obj)

            if not task_raw or cur_obj is None:
                continue

            sub_positions = [
                p for i, p in enumerate(range(1, 7), start=3)
                if i < len(row) and _truthy(row[i])
            ]
            cur_obj["tasks"].append({
                "name": task_raw,
                "due_date": _pdate(row[2]) if len(row) > 2 else None,
                "sub_obj_positions": sub_positions,
            })

    # ── Deliverables ──────────────────────────────────────────
    if "Deliverables" in wb.sheetnames:
        for row in wb["Deliverables"].iter_rows(min_row=2, values_only=True):
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if desc:
                preview["deliverables"].append(desc)

    # ── Forecasts ─────────────────────────────────────────────
    if "Forecasts" in wb.sheetnames:
        for row in wb["Forecasts"].iter_rows(min_row=2, values_only=True):
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if desc:
                preview["forecasts"].append(desc)

    # ── Risks ─────────────────────────────────────────────────
    if "Risks" in wb.sheetnames:
        for row in wb["Risks"].iter_rows(min_row=2, values_only=True):
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not desc:
                continue
            raw_rag = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "green"
            preview["risks"].append({
                "description": desc,
                "rag": raw_rag if raw_rag in _VALID_RAG else "green",
            })

    return preview


async def import_oppm_xlsx(
    session: AsyncSession,
    project_id: str,
    workspace_id: str,
    xlsx_bytes: bytes,
) -> dict:
    """Parse an OPPM import template .xlsx and append data to the project.

    Returns a summary dict with counts of created records.
    This is append-only — existing data is never deleted.
    Objectives are matched case-insensitively and reused if found.
    """

    def _truthy(v: object) -> bool:
        return str(v).strip().lower() in _TRUTHY if v else False

    def _pdate(v: object) -> _date | None:
        if v is None:
            return None
        if isinstance(v, _date):
            return v
        try:
            return _date.fromisoformat(str(v).strip()[:10])
        except ValueError:
            return None

    # Validate project belongs to this workspace
    prj_repo = ProjectRepository(session)
    prj = await prj_repo.find_by_id(project_id)
    if not prj or str(prj.workspace_id) != workspace_id:
        raise HTTPException(status_code=404, detail="Project not found in this workspace")

    # Parse workbook
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt XLSX file")

    summary = {
        "objectives_created": 0,
        "tasks_created": 0,
        "sub_objectives_created": 0,
        "deliverables_created": 0,
        "forecasts_created": 0,
        "risks_created": 0,
    }

    obj_r = ObjectiveRepository(session)
    sub_r = SubObjectiveRepository(session)
    dlv_r = DeliverableRepository(session)
    fct_r = ForecastRepository(session)
    rsk_r = RiskRepository(session)

    # ── Sub Objectives ────────────────────────────────────────
    if "Sub Objectives" in wb.sheetnames:
        ws = wb["Sub Objectives"]
        existing_sos = await sub_r.find_project_sub_objectives(project_id)
        used_positions = {so.position for so in existing_sos}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or (row[0] is None and (len(row) < 2 or row[1] is None)):
                break
            try:
                pos = int(row[0])
            except (TypeError, ValueError):
                continue
            label = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if 1 <= pos <= 6 and label and pos not in used_positions:
                await sub_r.create({"project_id": project_id, "position": pos, "label": label})
                used_positions.add(pos)
                summary["sub_objectives_created"] += 1

    # Build position → sub-obj id map for task linking
    all_sos = await sub_r.find_project_sub_objectives(project_id)
    so_by_pos = {so.position: so for so in all_sos}

    # ── Tasks ─────────────────────────────────────────────────
    if "Tasks" in wb.sheetnames:
        ws = wb["Tasks"]
        existing_objs = await obj_r.find_project_objectives(project_id)
        obj_by_title: dict[str, object] = {o.title.strip().lower(): o for o in existing_objs}
        cur_obj = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            # Stop at the first fully-blank row
            row_vals = row[:9] if len(row) >= 9 else row
            if all(v is None or str(v).strip() == "" for v in row_vals):
                break

            obj_raw = str(row[0]).strip() if row[0] else ""
            task_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            if obj_raw:
                key = obj_raw.lower()
                if key in obj_by_title:
                    cur_obj = obj_by_title[key]
                else:
                    new_o = await obj_r.create({
                        "project_id": project_id,
                        "title": obj_raw,
                        "sort_order": len(obj_by_title) + 1,
                    })
                    obj_by_title[key] = new_o
                    cur_obj = new_o
                    summary["objectives_created"] += 1

            if not task_raw or cur_obj is None:
                continue

            due = _pdate(row[2]) if len(row) > 2 else None
            linked_so_ids = [
                so_by_pos[p].id
                for i, p in enumerate(range(1, 7), start=3)
                if i < len(row) and _truthy(row[i]) and p in so_by_pos
            ]

            task = Task(
                id=_uuid.uuid4(),
                project_id=prj.id,
                oppm_objective_id=cur_obj.id,
                title=task_raw,
                status="todo",
                sort_order=summary["tasks_created"],
                due_date=due,
            )
            session.add(task)
            await session.flush()

            for so_id in linked_so_ids:
                session.add(TaskSubObjective(task_id=task.id, sub_objective_id=so_id))

            summary["tasks_created"] += 1

    # ── Deliverables ──────────────────────────────────────────
    if "Deliverables" in wb.sheetnames:
        ws = wb["Deliverables"]
        next_n = len(await dlv_r.find_project_deliverables(project_id)) + 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if desc:
                await dlv_r.create({"project_id": project_id, "item_number": next_n, "description": desc})
                next_n += 1
                summary["deliverables_created"] += 1

    # ── Forecasts ─────────────────────────────────────────────
    if "Forecasts" in wb.sheetnames:
        ws = wb["Forecasts"]
        next_n = len(await fct_r.find_project_forecasts(project_id)) + 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if desc:
                await fct_r.create({"project_id": project_id, "item_number": next_n, "description": desc})
                next_n += 1
                summary["forecasts_created"] += 1

    # ── Risks ─────────────────────────────────────────────────
    if "Risks" in wb.sheetnames:
        ws = wb["Risks"]
        next_n = len(await rsk_r.find_project_risks(project_id)) + 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            desc = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not desc:
                continue
            raw_rag = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "green"
            rag = raw_rag if raw_rag in _VALID_RAG else "green"
            await rsk_r.create({
                "project_id": project_id,
                "item_number": next_n,
                "description": desc,
                "rag": rag,
            })
            next_n += 1
            summary["risks_created"] += 1

    await session.commit()
    logger.info("OPPM import complete for project %s: %s", project_id, summary)
    return summary


async def import_oppm_json(
    session: AsyncSession,
    project_id: str,
    workspace_id: str,
    data: dict,
) -> dict:
    """Save structured OPPM data (as returned by the AI OCR extractor) to the DB.

    Accepts the same JSON shape returned by POST /api/v1/workspaces/{wid}/ai/oppm-extract:
    {
      "sub_objectives": [{"position": 1, "label": "..."}],
      "objectives": [
        {
          "title": "...",
          "tasks": [{"name": "...", "due_date": "YYYY-MM-DD|null", "sub_obj_positions": [1, 2]}]
        }
      ],
      "deliverables": ["..."],
      "forecasts": ["..."],
      "risks": [{"description": "...", "rag": "green|amber|red"}]
    }

    Import is append-only; objectives are matched by name (case-insensitive).
    """
    prj_repo = ProjectRepository(session)
    prj = await prj_repo.find_by_id(project_id)
    if not prj or str(prj.workspace_id) != workspace_id:
        raise HTTPException(status_code=404, detail="Project not found in this workspace")

    summary = {
        "objectives_created": 0,
        "tasks_created": 0,
        "sub_objectives_created": 0,
        "deliverables_created": 0,
        "forecasts_created": 0,
        "risks_created": 0,
    }

    obj_r = ObjectiveRepository(session)
    sub_r = SubObjectiveRepository(session)
    dlv_r = DeliverableRepository(session)
    fct_r = ForecastRepository(session)
    rsk_r = RiskRepository(session)

    # ── Sub Objectives ────────────────────────────────────────
    existing_sos = await sub_r.find_project_sub_objectives(project_id)
    used_positions = {so.position for so in existing_sos}
    for so in data.get("sub_objectives") or []:
        try:
            pos = int(so.get("position", 0))
        except (TypeError, ValueError):
            continue
        label = str(so.get("label", "")).strip()
        if 1 <= pos <= 6 and label and pos not in used_positions:
            await sub_r.create({"project_id": project_id, "position": pos, "label": label})
            used_positions.add(pos)
            summary["sub_objectives_created"] += 1

    all_sos = await sub_r.find_project_sub_objectives(project_id)
    so_by_pos = {so.position: so for so in all_sos}

    # ── Objectives + Tasks ────────────────────────────────────
    existing_objs = await obj_r.find_project_objectives(project_id)
    obj_by_title: dict[str, object] = {o.title.strip().lower(): o for o in existing_objs}

    for obj_data in data.get("objectives") or []:
        obj_title = str(obj_data.get("title", "")).strip()
        if not obj_title:
            continue

        key = obj_title.lower()
        if key in obj_by_title:
            cur_obj = obj_by_title[key]
        else:
            cur_obj = await obj_r.create({
                "project_id": project_id,
                "title": obj_title,
                "sort_order": len(obj_by_title) + 1,
            })
            obj_by_title[key] = cur_obj
            summary["objectives_created"] += 1

        for task_data in obj_data.get("tasks") or []:
            task_name = str(task_data.get("name", "")).strip()
            if not task_name:
                continue

            raw_due = task_data.get("due_date")
            due: _date | None = None
            if raw_due:
                try:
                    due = _date.fromisoformat(str(raw_due).strip()[:10])
                except ValueError:
                    pass

            linked_so_ids = [
                so_by_pos[p].id
                for p in (task_data.get("sub_obj_positions") or [])
                if isinstance(p, (int, float)) and int(p) in so_by_pos
            ]

            task = Task(
                id=_uuid.uuid4(),
                project_id=prj.id,
                oppm_objective_id=cur_obj.id,
                title=task_name,
                status="todo",
                sort_order=summary["tasks_created"],
                due_date=due,
            )
            session.add(task)
            await session.flush()

            for so_id in linked_so_ids:
                session.add(TaskSubObjective(task_id=task.id, sub_objective_id=so_id))

            summary["tasks_created"] += 1

    # ── Deliverables ──────────────────────────────────────────
    next_n = len(await dlv_r.find_project_deliverables(project_id)) + 1
    for desc in data.get("deliverables") or []:
        desc = str(desc).strip()
        if desc:
            await dlv_r.create({"project_id": project_id, "item_number": next_n, "description": desc})
            next_n += 1
            summary["deliverables_created"] += 1

    # ── Forecasts ─────────────────────────────────────────────
    next_n = len(await fct_r.find_project_forecasts(project_id)) + 1
    for desc in data.get("forecasts") or []:
        desc = str(desc).strip()
        if desc:
            await fct_r.create({"project_id": project_id, "item_number": next_n, "description": desc})
            next_n += 1
            summary["forecasts_created"] += 1

    # ── Risks ─────────────────────────────────────────────────
    next_n = len(await rsk_r.find_project_risks(project_id)) + 1
    for risk in data.get("risks") or []:
        desc = str(risk.get("description", "")).strip()
        if not desc:
            continue
        raw_rag = str(risk.get("rag", "green")).strip().lower()
        rag = raw_rag if raw_rag in _VALID_RAG else "green"
        await rsk_r.create({
            "project_id": project_id,
            "item_number": next_n,
            "description": desc,
            "rag": rag,
        })
        next_n += 1
        summary["risks_created"] += 1

    await session.commit()
    logger.info("OPPM JSON import complete for project %s: %s", project_id, summary)
    return summary
